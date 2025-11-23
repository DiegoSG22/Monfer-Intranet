# odontologia/views.py
import json 
from django.http import JsonResponse 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.forms import inlineformset_factory
from django.contrib import messages
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side 
from django.http import HttpResponse
# Modelos
from .models import Doctor, Atencion, DetalleAtencion, Examen
# Formularios
from .forms import AtencionForm, DetalleAtencionForm
from .forms import UserUpdateForm, DoctorProfileForm
from django.templatetags.static import static
import datetime 
from decimal import Decimal 

# --- Funciones Auxiliares ---

def get_saludo():
    now = timezone.localtime(timezone.now())
    hora_actual = now.hour
    if 5 <= hora_actual < 12: return "Buenos días"
    elif 12 <= hora_actual < 20: return "Buenas tardes"
    else: return "Buenas noches"

def get_doctor_data(user):
    if user.is_staff or user.is_superuser:
        return f"Admin. {user.first_name}", static('images/doctor_placeholder.png')

    nombre_doctor = user.first_name or user.username
    doctor_profile_pic_url = static('images/doctor_placeholder.png')
    try:
        doctor = user.doctor
        nombre_doctor = doctor.user.first_name or doctor.user.username
        if doctor.foto_perfil and hasattr(doctor.foto_perfil, 'url'):
            doctor_profile_pic_url = doctor.foto_perfil.url
    except Doctor.DoesNotExist:
        pass
    return nombre_doctor, doctor_profile_pic_url

# --- Vistas Principales ---

@login_required
def dashboard(request):
    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)
    es_admin = request.user.is_staff or request.user.is_superuser
    
    atenciones = []
    total_pacientes_hoy = 0
    atenciones_realizadas = 0
    ganancia_mensual = 0

    if es_admin:
        atenciones = Atencion.objects.all().order_by('-fecha', '-hora_atencion')
        now = timezone.localtime(timezone.now())
        total_pacientes_hoy = atenciones.filter(fecha=now.date()).count()
        atenciones_realizadas = atenciones.count()
    else:
        try:
            doctor = request.user.doctor
            atenciones = Atencion.objects.filter(doctor=doctor).order_by('-fecha', '-hora_atencion')
            now = timezone.localtime(timezone.now())
            total_pacientes_hoy = atenciones.filter(fecha=now.date()).count()
            atenciones_realizadas = atenciones.count()
        except Doctor.DoesNotExist:
            pass

    context = {
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'atenciones': atenciones,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin,
        'total_pacientes_hoy': total_pacientes_hoy,
        'atenciones_realizadas': atenciones_realizadas,
        'ganancia_mensual': ganancia_mensual,
    }
    return render(request, 'odontologia/dashboard.html', context)


@login_required
def registrar_atencion(request):
    try:
        doctor = request.user.doctor
    except Doctor.DoesNotExist:
        messages.error(request, 'Error: Solo los doctores pueden registrar atenciones.')
        return redirect('dashboard')

    DetalleFormSet = inlineformset_factory(
        Atencion, DetalleAtencion, form=DetalleAtencionForm, extra=1, can_delete=False
    )

    if request.method == 'POST':
        form = AtencionForm(request.POST)
        detalle_formset = DetalleFormSet(request.POST, prefix='detalles')

        if form.is_valid() and detalle_formset.is_valid():
            atencion = form.save(commit=False)
            atencion.doctor = doctor
            atencion.save()
            detalle_formset.instance = atencion
            detalle_formset.save()
            messages.success(request, '¡Atención guardada con éxito!')
            return redirect('dashboard')
        
        # CORRECCIÓN: Eliminamos el "else: messages.error(...)"
        # Si hay error, simplemente se re-renderiza la página y el HTML muestra los errores rojos.

    else:
        form = AtencionForm()
        detalle_formset = DetalleFormSet(prefix='detalles')

    nombre_doctor_display, doctor_profile_pic_display = get_doctor_data(request.user)
    saludo_display = get_saludo()

    context = {
        'form': form,
        'detalle_formset': detalle_formset,
        'nombre_doctor': nombre_doctor_display,
        'doctor_profile_pic': doctor_profile_pic_display,
        'saludo': saludo_display,
    }
    return render(request, 'odontologia/registrar_atencion.html', context)

# --- Vistas de Perfil ---

@login_required
def ver_perfil(request):
    try:
        doctor = request.user.doctor
    except Doctor.DoesNotExist:
        messages.warning(request, 'Esta vista es solo para doctores con perfil.')
        return redirect('dashboard')

    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)
    user_form = UserUpdateForm(instance=request.user)
    doctor_form = DoctorProfileForm(instance=doctor)

    context = {
        'user_form': user_form,
        'doctor_form': doctor_form,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
    }
    return render(request, 'odontologia/perfil.html', context)


@login_required
def editar_perfil(request):
    try:
        doctor = request.user.doctor
    except Doctor.DoesNotExist:
        messages.error(request, 'Perfil no encontrado.')
        return redirect('dashboard')

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        doctor_form = DoctorProfileForm(request.POST, request.FILES, instance=doctor)

        if user_form.is_valid() and doctor_form.is_valid():
            user_form.save()
            doctor_form.save()
            messages.success(request, '¡Perfil actualizado!')
            return redirect('ver_perfil')
        else:
            messages.error(request, 'Corrige los errores.')
            saludo = get_saludo()
            nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)
            context = {
                'user_form': user_form, 
                'doctor_form': doctor_form,
                'saludo': saludo, 
                'nombre_doctor': nombre_doctor, 
                'doctor_profile_pic': doctor_profile_pic
            }
            return render(request, 'odontologia/perfil.html', context)
    else:
        return redirect('ver_perfil')

# --- Vistas de Calendario, Detalles y Búsqueda ---

@login_required
def detalle_atencion(request, pk):
    es_admin = request.user.is_staff or request.user.is_superuser
    try:
        if es_admin:
            atencion = get_object_or_404(Atencion, pk=pk)
        else:
            doctor = request.user.doctor
            atencion = get_object_or_404(Atencion, pk=pk, doctor=doctor)
    except Doctor.DoesNotExist:
        messages.error(request, 'Acceso denegado.')
        return redirect('dashboard')
    except Atencion.DoesNotExist:
        messages.error(request, 'Atención no encontrada.')
        return redirect('dashboard')

    detalles_tratamiento = atencion.detalles.all()
    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)

    context = {
        'atencion': atencion,
        'detalles': detalles_tratamiento,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin,
    }
    return render(request, 'odontologia/detalle_atencion.html', context)

@login_required
def ver_calendario(request):
    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)
    es_admin = request.user.is_staff or request.user.is_superuser
    eventos_calendario = []
    
    if es_admin:
        atenciones = Atencion.objects.all()
    else:
        try:
            doctor = request.user.doctor
            atenciones = Atencion.objects.filter(doctor=doctor)
        except Doctor.DoesNotExist:
            atenciones = []

    for atencion in atenciones:
        start_datetime = datetime.datetime.combine(atencion.fecha, atencion.hora_atencion)
        titulo = f"{atencion.paciente_nombre} {atencion.paciente_apellido}"
        if es_admin:
            titulo = f"Dr. {atencion.doctor.user.last_name}: {titulo}"
            
        eventos_calendario.append({
            'id': atencion.pk,
            'title': titulo,
            'start': start_datetime.isoformat(),
            'allDay': False
        })

    context = {
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'atenciones_json': json.dumps(eventos_calendario),
        'es_admin': es_admin
    }
    return render(request, 'odontologia/calendario.html', context)

@login_required
def atencion_json(request, pk):
    es_admin = request.user.is_staff or request.user.is_superuser
    try:
        if es_admin:
            atencion = get_object_or_404(Atencion, pk=pk)
        else:
            atencion = get_object_or_404(Atencion, pk=pk, doctor=request.user.doctor)

        detalles = list(atencion.detalles.all().values('especialidad', 'descripcion', 'valor'))
        for d in detalles: d['valor'] = str(d['valor'])

        data = {
            'paciente': f"{atencion.paciente_nombre} {atencion.paciente_apellido}",
            'rut': atencion.paciente_rut,
            'fecha': atencion.fecha.strftime("%d/%m/%Y"),
            'hora': atencion.hora_atencion.strftime("%H:%M"),
            'motivo': atencion.motivo_visita or "No especificado",
            'pago': atencion.get_metodo_pago_display(),
            'detalles': detalles,
            'doctor': f"{atencion.doctor.user.first_name} {atencion.doctor.user.last_name}"
        }
        return JsonResponse(data)
    except (Doctor.DoesNotExist, Atencion.DoesNotExist):
        return JsonResponse({'error': 'No encontrado o no autorizado'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def editar_atencion(request, pk):
    es_admin = request.user.is_staff or request.user.is_superuser
    try:
        if es_admin:
            atencion = get_object_or_404(Atencion, pk=pk)
        else:
            doctor = request.user.doctor
            atencion = get_object_or_404(Atencion, pk=pk, doctor=doctor)
    except (Doctor.DoesNotExist, Atencion.DoesNotExist):
        messages.error(request, 'Permiso denegado.')
        return redirect('dashboard')

    DetalleFormSet = inlineformset_factory(
        Atencion, DetalleAtencion, form=DetalleAtencionForm, extra=1, can_delete=True
    )

    if request.method == 'POST':
        form = AtencionForm(request.POST, instance=atencion)
        detalle_formset = DetalleFormSet(request.POST, instance=atencion, prefix='detalles')

        if form.is_valid() and detalle_formset.is_valid():
            form.save()
            detalle_formset.save()
            messages.success(request, '¡Atención actualizada exitosamente!')
            return redirect('detalle_atencion', pk=atencion.pk)
        
        # CORRECCIÓN: Eliminamos el "else: messages.error(...)"

    else:
        form = AtencionForm(instance=atencion)
        detalle_formset = DetalleFormSet(instance=atencion, prefix='detalles')

    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)

    context = {
        'form': form,
        'detalle_formset': detalle_formset,
        'is_edit': True,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin
    }
    return render(request, 'odontologia/registrar_atencion.html', context)

@login_required
def eliminar_atencion(request, pk):
    es_admin = request.user.is_staff or request.user.is_superuser
    try:
        if es_admin:
            atencion = get_object_or_404(Atencion, pk=pk)
        else:
            doctor = request.user.doctor
            atencion = get_object_or_404(Atencion, pk=pk, doctor=doctor)
    except (Doctor.DoesNotExist, Atencion.DoesNotExist):
        messages.error(request, 'No tienes permiso para eliminar esta atención.')
        return redirect('dashboard')

    if request.method == 'POST':
        nombre = atencion.paciente_nombre
        atencion.delete()
        messages.success(request, f"La atención de {nombre} ha sido eliminada.")
        return redirect('dashboard')

    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)

    context = {
        'atencion': atencion,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
    }
    return render(request, 'odontologia/eliminar_atencion_confirm.html', context)

# --- Vistas de Gestión (Admin y Listados) ---

@login_required
def lista_doctores(request):
    es_admin = request.user.is_staff or request.user.is_superuser
    if not es_admin:
        messages.error(request, "Acceso restringido a administradores.")
        return redirect('dashboard')
    
    doctores = Doctor.objects.all()
    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)

    context = {
        'doctores': doctores,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin
    }
    return render(request, 'odontologia/lista_doctores.html', context)

@login_required
def atenciones_por_doctor(request, pk):
    """Vista detallada de un doctor específico con Buscador."""
    es_admin = request.user.is_staff or request.user.is_superuser
    if not es_admin:
        messages.error(request, "Acceso restringido.")
        return redirect('dashboard')

    doctor_objetivo = get_object_or_404(Doctor, pk=pk)
    atenciones = Atencion.objects.filter(doctor=doctor_objetivo).order_by('-fecha', '-hora_atencion')

    # Buscador por RUT
    query = request.GET.get('q')
    if query:
        atenciones = atenciones.filter(paciente_rut__icontains=query)

    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)

    context = {
        'doctor_objetivo': doctor_objetivo,
        'atenciones': atenciones,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin,
        'busqueda': query or ""
    }
    return render(request, 'odontologia/atenciones_doctor.html', context)

@login_required
def lista_atenciones(request):
    """Muestra TODAS las atenciones (según permiso) con buscador."""
    saludo = get_saludo()
    nombre_doctor, doctor_profile_pic = get_doctor_data(request.user)
    es_admin = request.user.is_staff or request.user.is_superuser

    if es_admin:
        atenciones = Atencion.objects.all().order_by('-fecha', '-hora_atencion')
    else:
        try:
            doctor = request.user.doctor
            atenciones = Atencion.objects.filter(doctor=doctor).order_by('-fecha', '-hora_atencion')
        except Doctor.DoesNotExist:
            atenciones = Atencion.objects.none()

    query = request.GET.get('q')
    if query:
        atenciones = atenciones.filter(paciente_rut__icontains=query)

    context = {
        'atenciones': atenciones,
        'saludo': saludo,
        'nombre_doctor': nombre_doctor,
        'doctor_profile_pic': doctor_profile_pic,
        'es_admin': es_admin,
        'busqueda': query or ""
    }
    return render(request, 'odontologia/lista_atenciones.html', context)

@login_required
def descargar_excel_doctor(request, pk):
    """Genera Excel profesional con ganancias."""
    es_admin = request.user.is_staff or request.user.is_superuser
    if not es_admin:
        messages.error(request, "No tienes permiso.")
        return redirect('dashboard')

    doctor = get_object_or_404(Doctor, pk=pk)
    atenciones = Atencion.objects.filter(doctor=doctor).order_by('-fecha', '-hora_atencion')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Reporte_{doctor.user.last_name}_{timezone.now().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Financiero"

    # Estilos
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    money_alignment = Alignment(horizontal='right')
    total_font = Font(name='Calibri', bold=True, size=12)
    total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    headers = ["Fecha", "Hora", "RUT Paciente", "Nombre Paciente", "Motivo", "Total Cobrado", "Ganancia Doctor (50%)"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    total_ganancias = Decimal('0.00')
    total_cobrado = Decimal('0.00')

    for atencion in atenciones:
        total_atencion = sum(detalle.valor for detalle in atencion.detalles.all())
        ganancia_doctor = total_atencion * Decimal('0.5')
        total_cobrado += total_atencion
        total_ganancias += ganancia_doctor

        ws.append([
            atencion.fecha.strftime("%d/%m/%Y"),
            atencion.hora_atencion.strftime("%H:%M"),
            atencion.paciente_rut,
            f"{atencion.paciente_nombre} {atencion.paciente_apellido}",
            atencion.motivo_visita,
            total_atencion,
            ganancia_doctor
        ])
        
        for col_num, cell in enumerate(ws[ws.max_row], 1):
            cell.border = thin_border
            if col_num >= 6:
                cell.number_format = '$ #,##0'
                cell.alignment = money_alignment

    # Fila Totales
    ws.append(["TOTALES", "", "", "", "", total_cobrado, total_ganancias])
    final_row = ws.max_row
    ws.merge_cells(start_row=final_row, start_column=1, end_row=final_row, end_column=5)
    
    for col_num, cell in enumerate(ws[final_row], 1):
        cell.font = total_font
        cell.fill = total_fill
        cell.border = thin_border
        if col_num == 1: cell.alignment = Alignment(horizontal='center')
        if col_num >= 6: 
            cell.number_format = '$ #,##0'
            cell.alignment = money_alignment

    # Ajustar ancho
    for col in ws.columns:
        max_len = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_len: max_len = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = (max_len + 2) * 1.2

    wb.save(response)
    return response